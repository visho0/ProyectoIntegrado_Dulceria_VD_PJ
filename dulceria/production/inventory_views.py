"""
Vistas para el módulo de inventario transaccional
"""
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.decorators.http import require_POST, require_http_methods
from django.utils import timezone
from datetime import datetime, date
from .models import MovimientoInventario, Bodega, Product, Proveedor
from .inventory_forms import MovimientoInventarioForm
from .views import get_user_role, get_pagination_per_page
from django.http import HttpResponse
from openpyxl import Workbook
from accounts.models import UserProfile


@login_required
def inventory_dashboard(request):
    """Dashboard del módulo de inventario con estadísticas"""
    role = get_user_role(request)
    
    # Solo usuarios autorizados pueden acceder
    if role not in ['admin', 'manager']:
        messages.error(request, 'No tienes permiso para acceder a esta página.')
        return redirect('dashboard')
    
    # Estadísticas del día
    hoy = timezone.now().date()
    movimientos_hoy = MovimientoInventario.objects.filter(fecha__date=hoy).count()
    
    # Stock total (suma de todos los productos)
    stock_total = Product.objects.filter(is_active=True, estado_aprobacion='APROBADO').aggregate(
        total=Sum('stock')
    )['total'] or 0
    
    # Productos únicos con stock
    productos_unicos = Product.objects.filter(
        is_active=True,
        estado_aprobacion='APROBADO',
        stock__gt=0
    ).count()
    
    # Últimos movimientos
    ultimos_movimientos = MovimientoInventario.objects.select_related(
        'producto', 'proveedor', 'bodega', 'creado_por'
    ).order_by('-fecha')[:10]
    
    context = {
        'movimientos_hoy': movimientos_hoy,
        'stock_total': int(stock_total),
        'productos_unicos': productos_unicos,
        'ultimos_movimientos': ultimos_movimientos,
        'user_role': role,
    }
    
    return render(request, 'production/inventory_dashboard.html', context)


@login_required
def movimientos_list(request):
    """Lista de movimientos de inventario con filtros y búsqueda"""
    role = get_user_role(request)
    
    # Admin, manager, employee (BODEGA) y viewer (CONSULTA) pueden ver
    # CONSULTA solo puede ver, no crear/editar (verificado en templates y vistas)
    if role not in ['admin', 'manager', 'employee', 'viewer']:
        messages.error(request, 'No tienes permiso para acceder a esta página.')
        return redirect('dashboard')
    
    # Obtener parámetros de búsqueda y filtros
    q = request.GET.get('q', '')
    tipo = request.GET.get('tipo', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    # Obtener movimientos - optimizado con select_related para evitar N+1 queries
    movimientos = MovimientoInventario.objects.select_related(
        'producto', 'producto__category',  # Optimizar acceso a producto y categoría
        'proveedor',  # Optimizar acceso a proveedor
        'bodega',  # Optimizar acceso a bodega
        'creado_por'  # Optimizar acceso a usuario creador
    ).all()
    
    # Aplicar búsqueda - optimizado para usar índices
    if q:
        search_conditions = Q()
        # Buscar en campos indexados primero
        search_conditions |= Q(producto__sku__icontains=q)  # Usa índice de producto.sku
        search_conditions |= Q(producto__name__icontains=q)  # Usa índice de producto.name
        search_conditions |= Q(proveedor__razon_social__icontains=q)  # Usa índice de proveedor.razon_social
        search_conditions |= Q(proveedor__rut__icontains=q)  # Usa índice de proveedor.rut
        search_conditions |= Q(doc_referencia__icontains=q)
        search_conditions |= Q(lote__icontains=q)
        search_conditions |= Q(serie__icontains=q)
        movimientos = movimientos.filter(search_conditions)
    
    # Filtrar por tipo - usa índice mov_tipo_fecha_idx
    if tipo:
        movimientos = movimientos.filter(tipo=tipo)
    
    # Filtrar por fecha - optimizado usando índice mov_fecha_idx
    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            movimientos = movimientos.filter(fecha__date__gte=fecha_desde_obj)
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            movimientos = movimientos.filter(fecha__date__lte=fecha_hasta_obj)
        except ValueError:
            pass
    
    # Ordenar por fecha descendente - usa índice mov_fecha_idx
    movimientos = movimientos.order_by('-fecha', '-created_at')
    
    # Paginación
    per_page = get_pagination_per_page(request, session_key='movimientos_per_page', default=25)
    paginator = Paginator(movimientos, per_page)
    page = request.GET.get('page', 1)
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    # Obtener tipos de movimiento para el filtro
    tipos_movimiento = MovimientoInventario.TIPO_MOVIMIENTO_CHOICES
    
    context = {
        'movimientos': page_obj,
        'q': q,
        'tipo': tipo,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'tipos_movimiento': tipos_movimiento,
        'per_page': per_page,
        'per_page_options': [25, 50, 100, 250, 500],  # Opciones optimizadas para grandes volúmenes
        'user_role': role,
    }
    
    return render(request, 'production/movimientos_list.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def movimiento_create(request):
    """Crear un nuevo movimiento de inventario"""
    role = get_user_role(request)
    
    # Solo admin, manager y employee (BODEGA) pueden crear movimientos
    # CONSULTA (viewer) NO puede crear movimientos
    if role not in ['admin', 'manager', 'employee']:
        messages.error(request, 'No tienes permiso para crear movimientos de inventario. Solo usuarios con permisos de administración, gerencia o bodega pueden realizar esta acción.')
        return redirect('movimientos_list')
    
    if request.method == 'POST':
        form = MovimientoInventarioForm(request.POST)
        if form.is_valid():
            movimiento = form.save(commit=False)
            movimiento.creado_por = request.user
            # Si no se especifica fecha, usar la actual
            if not movimiento.fecha:
                movimiento.fecha = timezone.now()
            movimiento.save()
            messages.success(request, f'Movimiento de {movimiento.get_tipo_display()} registrado exitosamente.')
            return redirect('movimientos_list')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = MovimientoInventarioForm(initial={'fecha': timezone.now()})
    
    context = {
        'form': form,
        'action': 'Registrar',
        'user_role': role,
    }
    
    return render(request, 'production/movimiento_form.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def movimiento_edit(request, pk):
    """Editar un movimiento de inventario existente"""
    role = get_user_role(request)
    
    # Solo admin y manager pueden editar movimientos
    # BODEGA (employee) y CONSULTA (viewer) NO pueden editar
    if role not in ['admin', 'manager']:
        messages.error(request, 'No tienes permiso para editar movimientos. Solo administradores y gerentes pueden realizar esta acción.')
        return redirect('movimientos_list')
    
    movimiento = get_object_or_404(MovimientoInventario, pk=pk)
    
    if request.method == 'POST':
        form = MovimientoInventarioForm(request.POST, instance=movimiento)
        if form.is_valid():
            # Guardar los campos readonly manualmente (no se envían en POST si están readonly)
            movimiento.fecha = form.cleaned_data.get('fecha', movimiento.fecha)
            movimiento.tipo = form.cleaned_data.get('tipo', movimiento.tipo)
            movimiento.producto = form.cleaned_data.get('producto', movimiento.producto)
            
            # Guardar el resto de los campos editables
            movimiento.proveedor = form.cleaned_data.get('proveedor')
            movimiento.bodega = form.cleaned_data.get('bodega')
            movimiento.cantidad = form.cleaned_data.get('cantidad')
            movimiento.lote = form.cleaned_data.get('lote', '')
            movimiento.serie = form.cleaned_data.get('serie', '')
            movimiento.fecha_vencimiento = form.cleaned_data.get('fecha_vencimiento')
            movimiento.doc_referencia = form.cleaned_data.get('doc_referencia', '')
            movimiento.observaciones = form.cleaned_data.get('observaciones', '')
            movimiento.motivo = form.cleaned_data.get('motivo', '')
            
            movimiento.save()
            messages.success(request, f'Movimiento actualizado exitosamente.')
            return redirect('movimientos_list')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = MovimientoInventarioForm(instance=movimiento)
    
    context = {
        'form': form,
        'movimiento': movimiento,
        'action': 'Editar',
        'user_role': role,
    }
    
    return render(request, 'production/movimiento_form.html', context)


@login_required
@require_POST
def movimiento_delete(request, pk):
    """Eliminar un movimiento de inventario"""
    role = get_user_role(request)
    
    # Solo admin y manager pueden eliminar
    if role not in ['admin', 'manager']:
        messages.error(request, 'No tienes permiso para eliminar movimientos.')
        return redirect('movimientos_list')
    
    movimiento = get_object_or_404(MovimientoInventario, pk=pk)
    movimiento_str = str(movimiento)
    
    movimiento.delete()
    
    messages.success(request, f'Movimiento "{movimiento_str}" eliminado exitosamente.')
    return redirect('movimientos_list')


@login_required
def export_inventory_excel(request):
    """Generar un archivo Excel con productos, usuarios por rol y movimientos"""
    role = get_user_role(request)

    if role not in ['admin', 'manager', 'employee']:
        messages.error(request, 'No tienes permiso para exportar la información de inventario.')
        return redirect('inventory_dashboard')

    wb = Workbook()

    # --- Hoja de productos ---
    ws_productos = wb.active
    ws_productos.title = 'Productos'
    ws_productos.append([
        'SKU', 'Nombre', 'Categoría', 'Estado Aprobación', 'Stock', 'Stock Mínimo',
        'Stock Máximo', 'Precio Venta', 'IVA (%)', 'Unidad Compra', 'Unidad Venta',
        'Factor Conversión', 'Perecible', 'Control Lote', 'Control Serie',
        'Fecha Vencimiento', 'Creado Por'
    ])

    productos = Product.objects.select_related('category', 'creado_por').all().order_by('sku')
    for producto in productos:
        categoria = producto.category.name if producto.category else ''
        creado_por = producto.creado_por.get_full_name() if producto.creado_por else ''
        if not creado_por and producto.creado_por:
            creado_por = producto.creado_por.username
        ws_productos.append([
            producto.sku,
            producto.name,
            categoria,
            producto.get_estado_aprobacion_display(),
            producto.stock,
            producto.stock_minimo,
            producto.stock_maximo or '',
            producto.price or '',
            producto.iva,
            producto.uom_compra,
            producto.uom_venta,
            producto.factor_conversion,
            'Sí' if producto.es_perecible else 'No',
            'Sí' if producto.control_por_lote else 'No',
            'Sí' if producto.control_por_serie else 'No',
            producto.fecha_vencimiento.strftime('%d-%m-%Y') if producto.fecha_vencimiento else '',
            creado_por,
        ])

    # --- Hojas de usuarios por rol ---
    perfiles = UserProfile.objects.select_related('user', 'organization').order_by('role', 'user__username')
    roles_definidos = dict(UserProfile.ROLE_CHOICES)

    for valor_rol, etiqueta in UserProfile.ROLE_CHOICES:
        perfiles_rol = [perfil for perfil in perfiles if perfil.role == valor_rol]
        if not perfiles_rol:
            continue

        nombre_hoja = f"Usuarios {etiqueta}"
        if len(nombre_hoja) > 31:
            nombre_hoja = nombre_hoja[:28] + '...'

        ws_rol = wb.create_sheet(title=nombre_hoja)
        ws_rol.append([
            'Username', 'Email', 'Nombres', 'Apellidos', 'Rol', 'Estado',
            'MFA Habilitado', 'Organización', 'Teléfono', 'Área / Unidad',
            'Observaciones', 'Último acceso'
        ])

        for perfil in perfiles_rol:
            usuario = perfil.user
            last_login = usuario.last_login
            if last_login and timezone.is_naive(last_login):
                last_login_display = last_login.strftime('%d-%m-%Y %H:%M')
            elif last_login:
                last_login_display = timezone.localtime(last_login).strftime('%d-%m-%Y %H:%M')
            else:
                last_login_display = ''

            ws_rol.append([
                usuario.username,
                usuario.email,
                usuario.first_name,
                usuario.last_name,
                roles_definidos.get(perfil.role, perfil.role),
                perfil.get_state_display(),
                'Sí' if perfil.mfa_enabled else 'No',
                perfil.organization.name if perfil.organization else '',
                perfil.phone,
                perfil.area,
                perfil.observaciones,
                last_login_display,
            ])

    # --- Hoja de movimientos ---
    ws_mov = wb.create_sheet(title='Movimientos')
    ws_mov.append([
        'Fecha', 'Tipo', 'SKU', 'Producto', 'Proveedor', 'Bodega', 'Cantidad',
        'Lote', 'Serie', 'Fecha Vencimiento', 'Doc. Referencia', 'Motivo',
        'Observaciones', 'Creado Por'
    ])

    movimientos = MovimientoInventario.objects.select_related(
        'producto', 'proveedor', 'bodega', 'creado_por'
    ).order_by('-fecha')

    for mov in movimientos:
        fecha_mov = mov.fecha
        if fecha_mov and timezone.is_naive(fecha_mov):
            fecha_mov_display = fecha_mov.strftime('%d-%m-%Y %H:%M')
        elif fecha_mov:
            fecha_mov_display = timezone.localtime(fecha_mov).strftime('%d-%m-%Y %H:%M')
        else:
            fecha_mov_display = ''

        creado_por = mov.creado_por.get_full_name() if mov.creado_por else ''
        if not creado_por and mov.creado_por:
            creado_por = mov.creado_por.username

        ws_mov.append([
            fecha_mov_display,
            mov.get_tipo_display(),
            mov.producto.sku if mov.producto else '',
            mov.producto.name if mov.producto else '',
            mov.proveedor.razon_social if mov.proveedor else '',
            mov.bodega.codigo if mov.bodega else '',
            mov.cantidad,
            mov.lote,
            mov.serie,
            mov.fecha_vencimiento.strftime('%d-%m-%Y') if mov.fecha_vencimiento else '',
            mov.doc_referencia,
            mov.motivo,
            mov.observaciones,
            creado_por,
        ])

    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    current_time = timezone.now()
    if timezone.is_naive(current_time):
        timestamp_str = current_time.strftime('%Y%m%d_%H%M%S')
    else:
        timestamp_str = timezone.localtime(current_time).strftime('%Y%m%d_%H%M%S')

    wb.save(response)
    response['Content-Disposition'] = f'attachment; filename="reporte_inventario_{timestamp_str}.xlsx"'
    return response

